#!/usr/bin/env python3
"""Add Market Landscape sheet to TANIA financial model from competitive analysis data."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

WB_PATH = "/Users/rossdevlin/Documents/rossdevl.in/TANIA_financial_model_v2.xlsx"

# Styles
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
TANIA_HIGHLIGHT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    if "Market Landscape" in wb.sheetnames:
        del wb["Market Landscape"]
    ws = wb.create_sheet("Market Landscape", index=len(wb.sheetnames))

    row = 1

    # --- Title ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="TANIA FAS  |  Market Landscape & Competitive Analysis")
    ws.cell(row=row, column=1).font = TITLE_FONT
    row += 2

    # --- Market context / potential capture ---
    ws.cell(row=row, column=1, value="1. MARKET CONTEXT & POTENTIAL CAPTURE")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1
    context_rows = [
        ("Global art market (2024)", "$57.5B", "12% YoY decline in high-end; lower/mid-tier resilient."),
        ("Target segment – dealers <$250K turnover", "+17%", "Aggregated sales growth (Tania's target market)."),
        ("Works <$50K as % of dealer sales", "85%", "Up from 73% in 2022 – mid-tier is growth segment."),
        ("Online dealer sales", "22%", "Up from 13% pre-pandemic; digital-first adoption."),
        ("Online sales to new buyers", "46%", "2024 – strong digital receptivity."),
        ("White-space", "Integrated software + physical", "No incumbent offers both; category-creating opportunity."),
        ("Key timing", "Gallery closures + post-imperial", "Displaced inventory + lean new galleries = ideal adopters."),
    ]
    ws.cell(row=row, column=1, value="Metric / Driver").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value / Signal").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Implication").font = Font(bold=True)
    row += 1
    for r in context_rows:
        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        row += 1
    row += 2

    # --- Software competitor comparison ---
    ws.cell(row=row, column=1, value="2. SOFTWARE COMPETITORS – FEATURE COMPARISON (Tania vs Gallery/Inventory Mgmt)")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1
    software_headers = ["Feature", "Artlogic / ArtCloud", "ARTERNAL", "ArtBinder", "Artwork Archive", "Others", "Tania FAS"]
    for c, val in enumerate(software_headers, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL if c > 1 else PatternFill()
        cell.border = THIN_BORDER
    row += 1
    software_matrix = [
        ("Inventory Mgmt", "Yes", "Yes", "Yes", "Yes", "Varies", "Yes"),
        ("CRM / Contacts", "Yes", "Core Focus", "Basic", "Basic", "Varies", "Planned"),
        ("Website Builder", "Yes", "No", "No", "Yes", "Some", "No"),
        ("Sales Pipeline", "Yes", "Yes", "Limited", "Limited", "No", "Planned"),
        ("Mobile/Device", "iOS App", "iPad/Web", "Yes", "Web", "Varies", "Yes (Built-in)"),
        ("Physical Storage", "No", "No", "No", "No", "No", "YES"),
        ("Shipping/Logistics", "No", "No", "No", "No", "No", "YES"),
        ("Crating/Handling", "No", "No", "No", "No", "No", "YES"),
        ("QR Code Tracking", "No", "No", "No", "No", "No", "YES"),
        ("Network Effects", "Marketplace", "Limited", "Limited", "Discovery", "No", "YES (Logistics)"),
    ]
    for r in software_matrix:
        for c, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            if c == 7 and str(val).upper() == "YES":
                cell.fill = TANIA_HIGHLIGHT
        row += 1
    row += 2

    # --- Physical storage / logistics comparison ---
    ws.cell(row=row, column=1, value="3. PHYSICAL STORAGE & LOGISTICS – COMPARISON (Tania vs Storage/Logistics)")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1
    phys_headers = ["Feature", "UOVO", "Hangman", "Mana Fine Arts", "Tania FAS"]
    for c, val in enumerate(phys_headers, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL if c > 1 else PatternFill()
        cell.border = THIN_BORDER
    row += 1
    phys_matrix = [
        ("Location(s)", "10+ US (NY, CA, CO, FL, TX, DE)", "Brooklyn, Hamptons, LA, Miami", "Jersey City, NJ", "Williamsburg, BK (+ planned upstate)"),
        ("Scale", "900K+ sq ft", "Multi-facility", "Single facility", "Early stage; 10% utilized"),
        ("Price Point", "Premium ($$$)", "Mid-Premium ($$)", "Mid-Range ($$)", "Affordable ($)"),
        ("Client Software", "Client Portal (basic)", "Internal only", "None", "Full Gallery SaaS"),
        ("Removal Fees", "Yes (punitive)", "Standard", "Standard", "Transparent"),
        ("White-Glove Service", "Yes (institutional)", "Yes (premium)", "Yes", "Yes (personalized)"),
        ("Hyperlocal NYC", "Bushwick + outer", "Brooklyn", "NJ (bridge/tunnel)", "Williamsburg (core art district)"),
    ]
    for r in phys_matrix:
        for c, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            if c == 5:
                cell.fill = TANIA_HIGHLIGHT
        row += 1
    row += 2

    # --- Competitive moat & takeaways ---
    ws.cell(row=row, column=1, value="4. COMPETITIVE MOAT & MARKET CAPTURE TAKEAWAYS")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1
    takeaways = [
        "Integration moat (Strong): No incumbent offers software + physical under one platform; replicating both is multi-year and capital-intensive.",
        "Cost moat (Moderate): Williamsburg + lean ops + software-subsidized storage enable undercutting UOVO while staying sustainable.",
        "Switching cost (Growing): Inventory on Tania + physical storage = dual friction to leave; pure-play cannot match.",
        "Market capture angle: Target displaced galleries (post-closure inventory), post-imperial galleries, and UOVO defectors; lead with software, monetize storage/logistics.",
        "Risk: Artlogic/ArtCloud merged (6,000+ clients, 15M artworks); window to establish integrated alternative is time-sensitive.",
    ]
    for t in takeaways:
        ws.cell(row=row, column=1, value=t)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Source: Tania FAS Competitive Landscape Analysis (Feb 2026).")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    wb.save(WB_PATH)
    print("Saved 'Market Landscape' sheet to", WB_PATH)

if __name__ == "__main__":
    main()
